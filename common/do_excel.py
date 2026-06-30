import openpyxl
import os


class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.method = None
        self.data = None
        self.expect = None
        self.type = None
        self.org_id = None
        self.city_code = None
        self.pc_data = None


class OperationExcel:
    def __init__(self, file_name):
        self.file_name = file_name
        self.excel = openpyxl.load_workbook(self.file_name)

    def read_data(self, sheet_name, row=2):
        sheet = self.excel[sheet_name]
        data_list = []
        for r in range(row, sheet.max_row + 1):
            obj = Case()
            obj.case_id = sheet.cell(r, 1).value
            obj.title = sheet.cell(r, 2).value
            obj.url = sheet.cell(r, 3).value
            obj.method = sheet.cell(r, 4).value
            raw_data = sheet.cell(r, 5).value
            obj.data = eval(raw_data) if isinstance(raw_data, str) else raw_data
            obj.expect = sheet.cell(r, 6).value
            data_list.append(obj)
        return data_list

    def write_data(self, sheet_name, row=1, col=1, value=None):
        sheet = self.excel[sheet_name]
        sheet.cell(row, col, value)
        self.excel.save(self.file_name)

    def get_sheet_names(self):
        return self.excel.sheetnames
